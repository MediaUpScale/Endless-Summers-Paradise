"""
VideoFactoryScheduler — Production Sync & PostPlanner Excel Generator
Video Factory Module

Usage:
    python scheduler_module.py --sync-and-generate-excel
    python scheduler_module.py --sync-and-generate-excel --offset 4h
    python scheduler_module.py --generate-excel-only --offset 4h
    python scheduler_module.py --list-library
"""

import os
import sys
import json
import argparse
import glob
import re
import subprocess
from datetime import datetime, timedelta, timezone
from pathlib import Path

# ── Timezone ───────────────────────────────────────────────────────────────────
try:
    from zoneinfo import ZoneInfo
    _TZ_NY = ZoneInfo("America/New_York")
except ImportError:
    _TZ_NY = timezone(timedelta(hours=-5))  # type: ignore[assignment]

# ── Auto-install dependencies ──────────────────────────────────────────────────
_REQUIRED = {
    "dotenv":   "python-dotenv",
    "boto3":    "boto3",
    "pandas":   "pandas",
    "openpyxl": "openpyxl",
}
for _mod, _pkg in _REQUIRED.items():
    try:
        __import__(_mod)
    except ImportError:
        print(f"[INSTALL] '{_pkg}' not found — installing...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", _pkg, "-q"])
        print(f"[OK] '{_pkg}' installed.")

import boto3
import openpyxl
from openpyxl.styles import Font, Alignment
from botocore.config import Config
from botocore.exceptions import ClientError
from dotenv import load_dotenv

load_dotenv(override=True)

# ── Directory structure ────────────────────────────────────────────────────────
FACTORY_OUTPUT    = Path("FACTORY_OUTPUT")
POSTPLANNER_DIR   = FACTORY_OUTPUT / "postplanner"

# Create directories on import so every path below is always valid
FACTORY_OUTPUT.mkdir(exist_ok=True)
POSTPLANNER_DIR.mkdir(exist_ok=True)

# ── Constants ──────────────────────────────────────────────────────────────────
MIN_DURATION_SECONDS  = 30
LIBRARY_FILE          = str(FACTORY_OUTPUT / "global_video_library.json")
SAMPLE_EXCEL          = "sample_bulk_posts_import.xlsx"
B2_PUBLIC_URL_BASE    = "https://MediaupscaleStorage.s3.us-east-005.backblazeb2.com"
DEFAULT_HASHTAGS      = ("#EndlessSummerParadise #JellyKingdom "
                         "#RetroFuturism #AIVideo #SurrealWorld #ImpossiblePlaces")


# ── Official Backblaze B2 resource factory ────────────────────────────────────
def get_b2_resource():
    """
    Official Backblaze B2 connection method (boto3.resource + SigV4).
    - endpoint_url  : hardcoded — immune to .env whitespace corruption
    - region_name   : must match the bucket region
    - signature_version='s3v4' : mandatory for B2
    - addressing_style='path'  : forces path-style URLs, required by B2
                                 (avoids boto3 defaulting to virtual-hosted
                                  style which B2 rejects on multipart uploads)
    """
    return boto3.resource(
        service_name="s3",
        endpoint_url="https://s3.us-east-005.backblazeb2.com",
        aws_access_key_id=os.getenv("B2_KEY_ID").strip(),
        aws_secret_access_key=os.getenv("B2_APPLICATION_KEY").strip(),
        region_name="us-east-005",
        config=Config(
            signature_version="s3v4",
            s3={"addressing_style": "path"},
        ),
    )


# ══════════════════════════════════════════════════════════════════════════════
class CaptionEngine:
    """
    Generates unique, contextual 3-line captions for each episode.

    Caches the result inside the episode's storyboard.json under:
        consolidated_metadata -> facebook_caption
    On subsequent runs the cached value is returned immediately — no
    regeneration, guaranteeing cross-platform consistency.

    Uniqueness is tracked via a (hook, sensory, cta) index combo set that
    is persisted in global_video_library.json under __caption_engine_state.
    """

    MAX_CHARS = 350

    # ── Template pools — Humble & Authentic / Scientific-Aesthetic ───────────
    # Tone: observational, curious, nostalgic. No brand names. No hard sells.
    _HOOKS = [
        "Something about {name} doesn't add up — and that's exactly the point.",
        "{name} looks like a memory from a summer that never existed.",
        "The physics of {name} are genuinely unexplained.",
        "{name} — where 1955 and the impossible meet.",
        "There's a specific kind of calm inside {name}.",
        "The geometry of {name} shouldn't hold. But it does.",
        "{name} moves the way dreams move — slowly, elastically, perfectly.",
        "Nobody modelled {name} on a computer. It just appeared.",
        "{name} has the light quality of something painted, not built.",
        "The silence inside {name} is the strangest part.",
        "Every angle of {name} looks like a still from a film that was never made.",
        "{name} behaves like it has memory.",
    ]

    _SENSORY = [
        "The medium is high-viscosity Jell-O. It moves in sheets, not splashes.",
        "Noon sunlight passes clean through every translucent wall.",
        "The whole structure vibrates at a frequency just below hearing.",
        "Surfaces yield underfoot, then spring back without a trace.",
        "Caustic light patterns shift across the jelly walls like slow fire.",
        "The water doesn't fall — it folds, layers, and resets.",
        "Everything here is see-through. You can watch the physics happen.",
        "The elastic deformation cycles are longer than you'd expect.",
        "Kodak Portra 400 colour cast over every surface — warm, slightly over-exposed.",
        "The 1950s silhouettes and the impossible geometry coexist without friction.",
        "Each impact sends a slow shockwave through the entire structure.",
        "Hard shadows at noon. Soft translucent glow from within.",
    ]

    _CLOSING = [
        "Filmed in one continuous movement.",
        "The architecture does the work.",
        "Shot on location in an impossible place.",
        "The material physics are not simulated.",
        "Every frame is vertical. Every surface is real.",
        "Jell-O behaves differently at this scale.",
        "The slide continues past the edge of the frame.",
        "The 1950s never looked this strange.",
        "You can see straight through to the other side.",
        "The light does something unusual near the top.",
        "The structure is still moving after the shot ends.",
        "Nothing was added in post.",
    ]

    def __init__(self, library: dict):
        state = library.get("__caption_engine_state", {})
        self._used: set[tuple] = {
            tuple(x) for x in state.get("used_combos", [])
        }

    def get_state(self) -> dict:
        """Return serialisable state for persistence in the library."""
        return {"used_combos": [list(x) for x in self._used]}

    # ── Public API ─────────────────────────────────────────────────────────────

    def get_or_generate(
        self,
        video_path: str,
        storyboard_data: "dict | list | None",
        storyboard_path: "Path | None",
    ) -> str:
        """
        Return the cached facebook_caption if already written to the
        storyboard, otherwise generate a unique one and cache it.
        """
        if isinstance(storyboard_data, dict):
            cached = (storyboard_data
                      .get("consolidated_metadata", {})
                      .get("facebook_caption", ""))
            if cached and cached.strip():
                return cached.strip()

        attraction = self._extract_attraction_name(video_path, storyboard_data)
        caption    = self._generate_unique(attraction)

        if storyboard_path and storyboard_path.exists():
            self._write_to_storyboard(storyboard_path, caption, storyboard_data)

        return caption

    # ── Helpers ────────────────────────────────────────────────────────────────

    @staticmethod
    def _extract_attraction_name(
        video_path: str,
        storyboard_data: "dict | list | None",
    ) -> str:
        """
        Parse the human-readable episode/attraction name.

        Priority:
          1. Episode folder name  (always specific to this video)
          2. project_title from SEO block with park prefix stripped
             — only used if it differs meaningfully from the folder name

        The folder name is preferred because project_title is often the
        shared park name ("The Jelly Kingdom") across all episodes.
        """
        # ── Folder name (primary — always episode-specific)
        folder = Path(video_path).parent.name
        folder = re.sub(r"_V\d+_LIVE$",    "", folder)
        folder = re.sub(r"_V\d+$",          "", folder)
        folder = re.sub(r"_\d{7,}$",        "", folder)
        folder = re.sub(r"_\d{8}_\d{6}$",   "", folder)
        folder_name = folder.replace("_", " ").replace("-", " ").strip()

        # ── SEO project_title (secondary — strip shared park prefix)
        if isinstance(storyboard_data, dict):
            seo   = storyboard_data.get("seo_metadata_usa_high_rpm", {}) or {}
            title = seo.get("project_title", "")
            for prefix in ("Endless Summer Paradise \u2014 ",
                           "Endless Summer Paradise - ",
                           "Endless Summer Paradise — "):
                if title.startswith(prefix):
                    title = title[len(prefix):]
                    break
            title = title.strip()
            # Only use title if it's more specific than the generic park name
            _generic = {"The Jelly Kingdom", "Jelly Kingdom", "Endless Summer Paradise"}
            if title and title not in _generic:
                return title

        return folder_name

    def _generate_unique(self, attraction_name: str) -> str:
        """
        Select a (hook, sensory, cta) combo that has not been used before.
        Selection is seeded from the attraction name for natural variety.
        Resets the used-set if all combinations are exhausted.
        """
        nh, ns, nc = len(self._HOOKS), len(self._SENSORY), len(self._CLOSING)
        seed = sum(ord(c) for c in attraction_name)

        for attempt in range(nh * ns * nc):
            # Strides must be coprime to pool sizes (all 12) so each dimension
            # cycles through ALL values before repeating:
            #   h stride = 1  → gcd(1,12)=1  ✓
            #   s stride = 5  → gcd(5,12)=1  ✓  (visits all 12 sensory lines)
            #   c stride = 11 → gcd(11,12)=1 ✓  (visits all 12 CTAs)
            h = (seed +  attempt)           % nh
            s = (seed + (attempt *  5) + 1) % ns
            c = (seed + (attempt * 11) + 2) % nc
            combo = (h, s, c)
            if combo in self._used:
                continue

            self._used.add(combo)
            hook    = self._HOOKS[h].format(name=attraction_name)
            sensory = self._SENSORY[s]
            cta     = self._CLOSING[c].format(name=attraction_name)
            caption = f"{hook}\n{sensory}\n{cta}"

            if len(caption) <= self.MAX_CHARS:
                return caption
            # Too long — shorten hook to bare name
            caption = f"{attraction_name}.\n{sensory}\n{cta}"
            if len(caption) <= self.MAX_CHARS:
                return caption

        # All combos exhausted — reset and recurse once
        self._used.clear()
        return self._generate_unique(attraction_name)

    @staticmethod
    def _write_to_storyboard(
        path: Path,
        caption: str,
        existing_data: "dict | list | None",
    ):
        """Write caption into storyboard.json under consolidated_metadata."""
        try:
            if existing_data is None:
                with open(path, encoding="utf-8", errors="replace") as f:
                    existing_data = json.loads(f.read())
            if not isinstance(existing_data, dict):
                return
            existing_data.setdefault("consolidated_metadata", {})
            existing_data["consolidated_metadata"]["facebook_caption"] = caption
            with open(path, "w", encoding="utf-8") as f:
                json.dump(existing_data, f, indent=2, ensure_ascii=False)
        except Exception as exc:
            print(f"  [WARN] Could not write consolidated_metadata "
                  f"to {path.name}: {exc}")


# ══════════════════════════════════════════════════════════════════════════════
class VideoFactoryScheduler:
    """
    Scans PRODUCTION_PATH for *_ULTIMATE_MASTER.mp4 files, uploads new ones
    to Backblaze B2, and generates a PostPlanner-compatible .xlsx schedule.
    """

    def __init__(self):
        self._validate_env()
        self.bucket_name     = os.getenv("B2_BUCKET_NAME", "MediaupscaleStorage").strip()
        self.production_path = self._resolve_production_path()
        self.b2              = get_b2_resource()
        self.bucket          = self.b2.Bucket(self.bucket_name)
        self.library: dict   = self._load_library()
        self.caption_engine  = CaptionEngine(self.library)

    # ── Startup ────────────────────────────────────────────────────────────────

    @staticmethod
    def _validate_env():
        required = ("B2_ENDPOINT_URL", "B2_KEY_ID",
                    "B2_APPLICATION_KEY", "B2_BUCKET_NAME")
        missing = [k for k in required if not os.getenv(k, "").strip()]
        if missing:
            raise EnvironmentError(
                f"Missing .env variables: {missing}\n"
                "Check your .env file and re-run."
            )

    @staticmethod
    def _resolve_production_path() -> str:
        raw = os.getenv("PRODUCTION_PATH", "").strip()
        for candidate in [raw, "Endless_Summers_Paradise - Production", "./production"]:
            if candidate and Path(candidate).exists():
                if candidate != raw:
                    print(f"[INFO] PRODUCTION_PATH '{raw}' not found — "
                          f"using '{candidate}' instead.")
                return candidate
        # Create the configured path as a last resort
        Path(raw).mkdir(parents=True, exist_ok=True)
        print(f"[INFO] Created production directory: {raw}")
        return raw

    # ── Library helpers ────────────────────────────────────────────────────────

    def _load_library(self) -> dict:
        if Path(LIBRARY_FILE).exists():
            with open(LIBRARY_FILE, encoding="utf-8") as f:
                return json.load(f)
        return {}

    def _save_library(self):
        # Persist caption engine state so uniqueness carries across runs
        if hasattr(self, "caption_engine"):
            self.library["__caption_engine_state"] = self.caption_engine.get_state()
        with open(LIBRARY_FILE, "w", encoding="utf-8") as f:
            json.dump(self.library, f, indent=2, ensure_ascii=False)
        print(f"[LIBRARY] Saved → {LIBRARY_FILE}")

    # ── Production scanner ─────────────────────────────────────────────────────

    def find_master_videos(self) -> list[str]:
        """Recursively find all *_ULTIMATE_MASTER.mp4 files."""
        pattern = os.path.join(self.production_path, "**", "*_ULTIMATE_MASTER.mp4")
        masters = sorted(glob.glob(pattern, recursive=True))

        if not masters:
            # Fallback: any .mp4 at subfolder level that isn't a raw scene clip
            _skip = {"_VIDEO", "_BAKED", "_soundscape", "_sfx", "_music"}
            all_mp4 = glob.glob(
                os.path.join(self.production_path, "**", "*.mp4"), recursive=True
            )
            masters = sorted(
                p for p in all_mp4
                if not any(tok in Path(p).name for tok in _skip)
                and Path(p).parent.name != Path(self.production_path).name
            )

        return masters

    # ── Duration check ─────────────────────────────────────────────────────────

    @staticmethod
    def get_duration(video_path: str) -> float:
        """Return video duration in seconds. Tries moviepy then OpenCV."""
        try:
            from moviepy import VideoFileClip
            with VideoFileClip(video_path) as clip:
                return clip.duration
        except Exception:
            pass
        try:
            from moviepy.editor import VideoFileClip
            with VideoFileClip(video_path) as clip:
                return clip.duration
        except Exception:
            pass
        try:
            import cv2
            cap = cv2.VideoCapture(video_path)
            fps    = cap.get(cv2.CAP_PROP_FPS) or 1
            frames = cap.get(cv2.CAP_PROP_FRAME_COUNT)
            cap.release()
            return frames / fps
        except Exception:
            pass
        print(f"  [WARN] Cannot read duration for {Path(video_path).name} — assuming valid.")
        return MIN_DURATION_SECONDS + 1

    # ── Metadata extraction ────────────────────────────────────────────────────

    def load_metadata(self, video_path: str) -> dict:
        """
        Build the full metadata dict for a video.

        Caption strategy (in priority order):
          1. storyboard.json → consolidated_metadata.facebook_caption  (cached)
          2. Generate a unique contextual caption via CaptionEngine and
             write it back into storyboard.json for future consistency.
          3. Filename-based fallback if no storyboard found at all.

        Hashtags come from the SEO block; pin_title from project_title.
        """
        folder = Path(video_path).parent
        stem   = Path(video_path).stem

        storyboard_data: dict | list | None = None
        storyboard_path: Path | None        = None

        for candidate in ("storyboard.json", "production_script.json"):
            jp = folder / candidate
            if not jp.exists():
                continue
            try:
                with open(jp, encoding="utf-8", errors="replace") as f:
                    storyboard_data = json.loads(f.read())
                storyboard_path = jp
                break
            except Exception as exc:
                print(f"  [WARN] Could not parse {jp.name}: {exc}")

        # Generate / retrieve unique contextual caption
        caption = self.caption_engine.get_or_generate(
            video_path, storyboard_data, storyboard_path
        )

        # Pull hashtags and pin_title from SEO block (independent of caption)
        hashtags  = DEFAULT_HASHTAGS
        pin_title = stem.replace("_", " ").replace("-", " ")
        if isinstance(storyboard_data, dict):
            seo       = storyboard_data.get("seo_metadata_usa_high_rpm", {}) or storyboard_data
            instagram = seo.get("instagram", {}) or {}
            facebook  = seo.get("facebook_meta", {}) or {}
            pinterest = seo.get("pinterest", {}) or {}
            hashtags  = (instagram.get("hashtags")
                         or facebook.get("hashtags")
                         or DEFAULT_HASHTAGS)
            pin_title = (pinterest.get("pin_title")
                         or seo.get("project_title")
                         or pin_title)

        return {
            "caption":   caption,
            "hashtags":  hashtags,
            "full_text": f"{caption}\n\n{hashtags}".strip(),
            "pin_title": pin_title,
        }

    @staticmethod
    def _parse_seo(data: dict | list, stem: str) -> dict:
        readable = stem.replace("_", " ").replace("-", " ")
        _fallback = {
            "caption":   readable,
            "hashtags":  DEFAULT_HASHTAGS,
            "full_text": f"{readable}\n\n{DEFAULT_HASHTAGS}",
            "pin_title": readable,
        }

        if isinstance(data, list):
            return _fallback

        seo       = data.get("seo_metadata_usa_high_rpm") or data
        instagram = seo.get("instagram") or {}
        facebook  = seo.get("facebook_meta") or {}
        pinterest = seo.get("pinterest") or {}

        caption   = (instagram.get("caption") or facebook.get("caption") or readable)
        hashtags  = (instagram.get("hashtags") or facebook.get("hashtags") or DEFAULT_HASHTAGS)
        pin_title = (pinterest.get("pin_title") or seo.get("project_title") or readable)

        return {
            "caption":   caption,
            "hashtags":  hashtags,
            "full_text": f"{caption}\n\n{hashtags}".strip(),
            "pin_title": pin_title,
        }

    # ── B2 upload (boto3.resource pattern) ────────────────────────────────────

    def _b2_object_exists(self, key: str) -> bool:
        """
        Check existence via load(). Returns False for 403/404 (B2 returns 403
        for missing objects on private buckets), re-raises anything else.
        """
        try:
            self.b2.Object(self.bucket_name, key).load()
            return True
        except ClientError as e:
            code = e.response["Error"]["Code"]
            if code in ("404", "NoSuchKey"):
                return False
            if code == "403":
                print(f"  [WARN] HeadObject 403 for '{key}' — "
                      "treating as not-yet-uploaded, proceeding.")
                return False
            raise

    def upload_to_b2(self, local_path: str) -> str:
        """
        Upload local_path to B2 using boto3.resource.
        The B2 object key is always the bare filename (no folder nesting).
        Returns 'Success' or 'Already Uploaded'.
        """
        key = Path(local_path).name  # flat key — never nested

        if self._b2_object_exists(key):
            return "Already Uploaded"

        size_mb = os.path.getsize(local_path) / (1024 * 1024)
        print(f"  [UPLOAD] {key} ({size_mb:.1f} MB) → B2:{self.bucket_name} …")

        self.b2.Object(self.bucket_name, key).upload_file(
            local_path,
            ExtraArgs={"ContentType": "video/mp4"},
        )
        return "Success"

    @staticmethod
    def get_public_url(filename: str) -> str:
        return f"{B2_PUBLIC_URL_BASE}/{Path(filename).name}"

    # ── Offset parser ─────────────────────────────────────────────────────────

    @staticmethod
    def _parse_offset(offset_str: str) -> timedelta:
        m = re.fullmatch(r"(\d+(?:\.\d+)?)\s*([hHmM])", offset_str.strip())
        if not m:
            raise ValueError(
                f"Invalid offset '{offset_str}'. Use e.g. '4h' or '30m'."
            )
        value, unit = float(m.group(1)), m.group(2).lower()
        return timedelta(hours=value) if unit == "h" else timedelta(minutes=value)

    # ── Main sync pipeline ─────────────────────────────────────────────────────

    def sync(self) -> list[dict]:
        """
        Scan PRODUCTION_PATH, filter by duration, upload new files to B2,
        and update global_video_library.json.
        Preserves is_posted status for videos already in the library.
        Returns the full list of processed records.
        """
        print(f"\n[SCAN] Production path: {self.production_path}")
        videos = self.find_master_videos()
        print(f"[SCAN] Found {len(videos)} master video(s).\n")

        processed = []

        for vpath in videos:
            name = Path(vpath).name
            print(f"─── {name}")

            # Duration gate
            duration = self.get_duration(vpath)
            if duration <= MIN_DURATION_SECONDS:
                print(f"  [SKIP] {duration:.1f}s ≤ {MIN_DURATION_SECONDS}s")
                continue
            print(f"  [OK]   Duration: {duration:.1f}s")

            # Metadata
            meta = self.load_metadata(vpath)

            # Upload
            try:
                status = self.upload_to_b2(vpath)
            except ClientError as exc:
                code = exc.response["Error"]["Code"]
                msg  = exc.response["Error"].get("Message", "")
                print(f"  [B2 ERROR] {code}: {msg} — skipping.")
                continue

            url = self.get_public_url(name)
            print(f"  [{status.upper()}] {url}")

            # Preserve is_posted if record already exists
            existing  = self.library.get(name, {})
            is_posted = existing.get("is_posted", False)

            self.library[name] = {
                "filename":   name,
                "local_path": str(Path(vpath).resolve()),
                "b2_url":     url,
                "duration_s": round(duration, 2),
                "status":     status,
                "caption":    meta["caption"],
                "hashtags":   meta["hashtags"],
                "full_text":  meta["full_text"],
                "pin_title":  meta["pin_title"],
                "is_posted":  is_posted,
                "synced_at":  datetime.now(timezone.utc).isoformat(),
            }
            processed.append(self.library[name])

        self._save_library()
        print(f"\n[SYNC] Done — {len(processed)} video(s) processed.")
        return processed

    # ── Excel generation ───────────────────────────────────────────────────────

    def generate_excel(
        self,
        records: list[dict] | None = None,
        offset: str | None = None,
    ):
        """
        Generate a PostPlanner .xlsx that clones the 4 header rows from
        sample_bulk_posts_import.xlsx, then appends one data row per
        unposted video.

        offset=None  → Queue mode (POSTING TIME blank)
        offset='4h'  → Scheduled mode starting from now (America/New_York),
                        each subsequent row staggered by the interval.
                        Format: MM/DD/YYYY HH:MM
        """
        if records is None:
            records = list(self.library.values())

        unposted = [r for r in records if not r.get("is_posted", False)]
        if not unposted:
            print("[EXCEL] All videos are already marked is_posted=true — nothing to export.")
            return

        timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path  = POSTPLANNER_DIR / f"PostPlanner_Export_{timestamp}.xlsx"
        print(f"\n[EXCEL] Building {excel_path.name} — {len(unposted)} unposted video(s)…")

        # Scheduling setup
        offset_delta:   timedelta | None = None
        schedule_start: datetime | None  = None
        if offset:
            offset_delta   = self._parse_offset(offset)
            schedule_start = datetime.now(_TZ_NY)
            print(f"  [SCHEDULE] Start {schedule_start.strftime('%m/%d/%Y %H:%M')} NY  "
                  f"| Interval {offset}")

        # ── Clone header rows from sample file
        if Path(SAMPLE_EXCEL).exists():
            src_wb = openpyxl.load_workbook(SAMPLE_EXCEL)
            src_ws = src_wb.active

            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = src_ws.title

            # Copy column widths
            for i in range(1, 7):
                col = openpyxl.utils.get_column_letter(i)
                if col in src_ws.column_dimensions:
                    new_ws.column_dimensions[col].width = \
                        src_ws.column_dimensions[col].width

            # Copy first 4 header rows verbatim (values + styles)
            for r in range(1, 5):
                for c in range(1, 7):
                    src  = src_ws.cell(row=r, column=c)
                    dest = new_ws.cell(row=r, column=c, value=src.value)
                    if src.font:      dest.font      = src.font.copy()
                    if src.fill:      dest.fill      = src.fill.copy()
                    if src.alignment: dest.alignment = src.alignment.copy()
        else:
            print(f"  [WARN] '{SAMPLE_EXCEL}' not found — using minimal header.")
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = "Sheet1"
            new_ws.append([
                "## Comment row — ignored on upload.",
                "BULK UPLOAD VERSION 2 — DO NOT EDIT this cell.",
                None, None, None, None,
            ])
            hdr = ["## POSTING TIME", "CAPTION", "CONTENT: LINK",
                   "CONTENT: MEDIA", "POST TYPE", "PIN TITLE"]
            new_ws.append(hdr)
            for cell in new_ws[2]:
                cell.font = Font(bold=True)

        # ── Data rows (start at row 5)
        for i, rec in enumerate(unposted):
            row = 5 + i

            if offset_delta and schedule_start:
                post_time = (schedule_start + offset_delta * i).strftime("%m/%d/%Y %H:%M")
            else:
                post_time = None   # blank = Queue mode

            caption = rec.get("full_text") or rec.get("caption", "")
            url     = rec.get("b2_url", "")

            new_ws.cell(row=row, column=1, value=post_time)   # POSTING TIME
            new_ws.cell(row=row, column=2, value=caption)     # CAPTION
            new_ws.cell(row=row, column=3, value=None)        # CONTENT: LINK  (empty)
            new_ws.cell(row=row, column=4, value=url)         # CONTENT: MEDIA
            new_ws.cell(row=row, column=5, value="VIDEO")     # POST TYPE
            new_ws.cell(row=row, column=6, value=None)        # PIN TITLE      (empty)

            new_ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

        new_wb.save(excel_path)
        print(f"[EXCEL] Saved → {excel_path}  ({len(unposted)} rows)")

        # ── Mark exported videos as posted
        for rec in unposted:
            if rec["filename"] in self.library:
                self.library[rec["filename"]]["is_posted"] = True
        self._save_library()
        print(f"[LIBRARY] {len(unposted)} video(s) marked is_posted=true")

    # ── Orchestrator ───────────────────────────────────────────────────────────

    def run_sync_and_generate(self, offset: str | None = None):
        records = self.sync()
        self.generate_excel(records, offset=offset)
        print(f"\n[DONE]\n  Library : {LIBRARY_FILE}\n  Schedule: {POSTPLANNER_DIR}/")


# ── CLI ────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Mediaupscale Video Factory — B2 Sync & PostPlanner Scheduler",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "--sync-and-generate-excel",
        action="store_true",
        help="Scan production folder, upload new videos to B2, generate .xlsx",
    )
    parser.add_argument(
        "--generate-excel-only",
        action="store_true",
        help="Re-generate .xlsx from existing library (no upload)",
    )
    parser.add_argument(
        "--offset",
        metavar="INTERVAL",
        default=None,
        help=(
            "Scheduling interval (e.g. '4h', '30m').\n"
            "Omit = Queue mode (blank POSTING TIME).\n"
            "Provided = staggered from now in America/New_York."
        ),
    )
    parser.add_argument(
        "--list-library",
        action="store_true",
        help="Print all entries in global_video_library.json",
    )

    args = parser.parse_args()
    if not any([args.sync_and_generate_excel,
                args.generate_excel_only,
                args.list_library]):
        parser.print_help()
        return

    scheduler = VideoFactoryScheduler()

    if args.sync_and_generate_excel:
        scheduler.run_sync_and_generate(offset=args.offset)

    elif args.generate_excel_only:
        scheduler.generate_excel(offset=args.offset)

    elif args.list_library:
        lib = scheduler.library
        if not lib:
            print("[LIBRARY] Empty — run --sync-and-generate-excel first.")
            return
        print(f"\n{'FILENAME':<58} {'POSTED':<8} {'DURATION':>10}")
        print("─" * 80)
        for rec in lib.values():
            posted   = "✓" if rec.get("is_posted") else "✗"
            duration = f"{rec.get('duration_s', '?')}s"
            print(f"  {rec['filename']:<56} [{posted}]  {duration:>9}")
        print(f"\nTotal: {len(lib)} video(s)")


if __name__ == "__main__":
    main()
